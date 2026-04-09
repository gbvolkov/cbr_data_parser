from __future__ import annotations

import csv
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


METRIC_START_COLUMN = 4
METRIC_END_COLUMN = 16
EXPECTED_METRIC_COUNT = 13
HIERARCHY_SEPARATOR = " > "
SECTION_HEADER_RE = re.compile(r"^[IVXLCDM]+\.")
TOTAL_LABEL_RE = re.compile(r"[–-]\s*всего$")


def validate_layout(worksheet) -> None:
    if normalize_text(worksheet["A5"].value) != "Вид страхования":
        raise ValueError("Expected cell A5 to contain 'Вид страхования'.")

    if worksheet.max_column != 17:
        raise ValueError(f"Expected 17 columns, got {worksheet.max_column}.")


def extract_report_metadata(worksheet) -> dict[str, str]:
    report_date_cell = normalize_text(worksheet["A3"].value)
    report_period_cell = normalize_text(worksheet["A4"].value)

    if not report_date_cell.startswith("Дата составления отчета: "):
        raise ValueError("Unexpected report date cell format in A3.")
    if not report_period_cell.startswith("Отчетный период: "):
        raise ValueError("Unexpected report period cell format in A4.")

    report_date = report_date_cell.removeprefix("Дата составления отчета: ").strip()
    report_period = report_period_cell.removeprefix("Отчетный период: ").strip()
    report_date_iso = datetime.strptime(report_date, "%d.%m.%Y").date().isoformat()

    return {
        "report_date": report_date,
        "report_date_iso": report_date_iso,
        "report_period": report_period,
    }


def build_merged_header_lookup(worksheet) -> dict[tuple[int, int], str]:
    merged_header_values: dict[tuple[int, int], str] = {}

    for merged_range in worksheet.merged_cells.ranges:
        min_column, min_row, max_column, max_row = merged_range.bounds
        if max_row < 5 or min_row > 7:
            continue

        anchor_value = normalize_text(worksheet.cell(min_row, min_column).value)
        for row_index in range(min_row, max_row + 1):
            for column_index in range(min_column, max_column + 1):
                merged_header_values[(row_index, column_index)] = anchor_value

    return merged_header_values


def build_merged_value_lookup(
    worksheet,
    row_start: int | None = None,
    row_end: int | None = None,
    column_start: int | None = None,
    column_end: int | None = None,
) -> dict[tuple[int, int], str]:
    merged_values: dict[tuple[int, int], str] = {}

    for merged_range in worksheet.merged_cells.ranges:
        min_column, min_row, max_column, max_row = merged_range.bounds
        if row_start is not None and max_row < row_start:
            continue
        if row_end is not None and min_row > row_end:
            continue
        if column_start is not None and max_column < column_start:
            continue
        if column_end is not None and min_column > column_end:
            continue

        actual_row_start = min_row if row_start is None else max(min_row, row_start)
        actual_row_end = max_row if row_end is None else min(max_row, row_end)
        actual_column_start = min_column if column_start is None else max(min_column, column_start)
        actual_column_end = max_column if column_end is None else min(max_column, column_end)

        anchor_value = normalize_text(worksheet.cell(min_row, min_column).value)
        for row_index in range(actual_row_start, actual_row_end + 1):
            for column_index in range(actual_column_start, actual_column_end + 1):
                merged_values[(row_index, column_index)] = anchor_value

    return merged_values


def extract_metric_labels(
    worksheet,
    merged_header_values: dict[tuple[int, int], str],
) -> list[str]:
    metric_labels: list[str] = []

    for column_index in range(METRIC_START_COLUMN, METRIC_END_COLUMN + 1):
        parts: list[str] = []
        for row_index in range(5, 8):
            value = normalize_text(worksheet.cell(row_index, column_index).value)
            if not value:
                value = merged_header_values.get((row_index, column_index), "")
            if value and (not parts or parts[-1] != value):
                parts.append(value)

        if not parts:
            raise ValueError(f"Metric header is empty in column {get_column_letter(column_index)}.")

        metric_labels.append(HIERARCHY_SEPARATOR.join(parts))

    if len(metric_labels) != EXPECTED_METRIC_COUNT:
        raise ValueError(
            f"Expected {EXPECTED_METRIC_COUNT} metric columns, got {len(metric_labels)}."
        )

    return metric_labels


def find_metric_column_indexes(
    worksheet,
    header_row_start: int,
    header_row_end: int,
    metric_start_column: int,
    metric_end_column: int | None = None,
    merged_values: dict[tuple[int, int], str] | None = None,
) -> list[int]:
    resolved_metric_end_column = worksheet.max_column if metric_end_column is None else metric_end_column
    metric_column_indexes: list[int] = []

    for column_index in range(metric_start_column, resolved_metric_end_column + 1):
        has_header_value = any(
            resolve_cell_text(worksheet, row_index, column_index, merged_values)
            for row_index in range(header_row_start, header_row_end + 1)
        )
        if has_header_value:
            metric_column_indexes.append(column_index)

    return metric_column_indexes


def extract_metric_labels_for_columns(
    worksheet,
    column_indexes: list[int],
    header_row_start: int,
    header_row_end: int,
    merged_values: dict[tuple[int, int], str] | None = None,
) -> list[str]:
    metric_labels: list[str] = []

    for column_index in column_indexes:
        parts: list[str] = []
        for row_index in range(header_row_start, header_row_end + 1):
            value = resolve_cell_text(worksheet, row_index, column_index, merged_values)
            if value and (not parts or parts[-1] != value):
                parts.append(value)

        if not parts:
            raise ValueError(f"Metric header is empty in column {get_column_letter(column_index)}.")

        metric_labels.append(HIERARCHY_SEPARATOR.join(parts))

    return metric_labels


def find_data_row_indexes(worksheet) -> list[int]:
    data_row_indexes: list[int] = []

    for row_index in range(8, worksheet.max_row + 1):
        has_metric_value = any(
            worksheet.cell(row_index, column_index).value is not None
            for column_index in range(METRIC_START_COLUMN, METRIC_END_COLUMN + 1)
        )
        if has_metric_value:
            data_row_indexes.append(row_index)

    if not data_row_indexes:
        raise ValueError("No data rows found in the worksheet.")

    return data_row_indexes


def row_has_any_metric_value(worksheet, row_index: int, metric_column_indexes: list[int]) -> bool:
    return any(worksheet.cell(row_index, column_index).value is not None for column_index in metric_column_indexes)


def extract_insurance_columns(worksheet, data_row_indexes: list[int]) -> list[dict]:
    insurance_columns: list[dict] = []
    current_section_label: str | None = None
    current_a_path: list[str] = []
    current_b_path: list[str] | None = None
    previous_label_column: int | None = None
    previous_label: str | None = None

    for row_index in data_row_indexes:
        header_cell_column, source_label = extract_row_label(worksheet, row_index)
        values = [
            worksheet.cell(row_index, column_index).value
            for column_index in range(METRIC_START_COLUMN, METRIC_END_COLUMN + 1)
        ]

        if header_cell_column == 1:
            current_b_path = None

            if is_section_header(source_label):
                current_section_label = source_label
                current_a_path = [source_label]
            elif current_section_label is None:
                current_a_path = [source_label]
            elif is_total_label(source_label):
                current_a_path = [current_section_label, source_label]
            elif previous_label_column in (2, 3):
                if len(current_a_path) < 2:
                    raise ValueError(
                        f"Cannot place row {row_index} after nested row without A-path context."
                    )
                current_a_path = [*current_a_path[:-1], source_label]
            elif previous_label_column == 1 and previous_label is not None:
                if is_section_header(previous_label):
                    current_a_path = [current_section_label, source_label]
                elif is_total_label(previous_label):
                    current_a_path = [*current_a_path, source_label]
                else:
                    if len(current_a_path) < 2:
                        raise ValueError(
                            f"Cannot place row {row_index} as A-level sibling without parent context."
                        )
                    current_a_path = [*current_a_path[:-1], source_label]
            else:
                raise ValueError(f"Cannot resolve hierarchy for row {row_index}.")

            path = current_a_path
        elif header_cell_column == 2:
            if not current_a_path:
                raise ValueError(f"Row {row_index} has B-level label without A-level context.")
            current_b_path = [*current_a_path, source_label]
            path = current_b_path
        elif header_cell_column == 3:
            if current_b_path is None:
                raise ValueError(f"Row {row_index} has C-level label without B-level context.")
            path = [*current_b_path, source_label]
        else:
            raise ValueError(f"Unexpected header cell column {header_cell_column} in row {row_index}.")

        insurance_columns.append(
            {
                "excel_row": row_index,
                "header_cell_column": header_cell_column,
                "source_label": source_label,
                "path": path,
                "column_name": HIERARCHY_SEPARATOR.join(path),
                "values": values,
            }
        )

        previous_label_column = header_cell_column
        previous_label = source_label

    return insurance_columns


def extract_row_label(worksheet, row_index: int) -> tuple[int, str]:
    found_cells: list[tuple[int, str]] = []

    for column_index in range(1, 4):
        value = normalize_text(worksheet.cell(row_index, column_index).value)
        if value:
            found_cells.append((column_index, value))

    if len(found_cells) != 1:
        raise ValueError(
            f"Expected exactly one row-header cell in A:C for row {row_index}, got {found_cells}."
        )

    return found_cells[0]


def extract_row_label_for_columns(
    worksheet,
    row_index: int,
    header_start_column: int,
    header_end_column: int,
) -> tuple[int, str]:
    found_cells: list[tuple[int, str]] = []

    for column_index in range(header_start_column, header_end_column + 1):
        value = normalize_text(worksheet.cell(row_index, column_index).value)
        if value:
            found_cells.append((column_index, value))

    if len(found_cells) != 1:
        raise ValueError(
            "Expected exactly one row-header cell "
            f"in {get_column_letter(header_start_column)}:{get_column_letter(header_end_column)} "
            f"for row {row_index}, got {found_cells}."
        )

    return found_cells[0]


def extract_row_path(
    worksheet,
    row_index: int,
    header_start_column: int,
    header_end_column: int,
    merged_values: dict[tuple[int, int], str] | None = None,
) -> list[str]:
    path: list[str] = []

    for column_index in range(header_start_column, header_end_column + 1):
        value = resolve_cell_text(worksheet, row_index, column_index, merged_values)
        if value and (not path or path[-1] != value):
            path.append(value)

    return path


def resolve_cell_text(
    worksheet,
    row_index: int,
    column_index: int,
    merged_values: dict[tuple[int, int], str] | None = None,
) -> str:
    value = normalize_text(worksheet.cell(row_index, column_index).value)
    if value:
        return value
    if merged_values is None:
        return ""
    return merged_values.get((row_index, column_index), "")


def normalize_text(value) -> str:
    if value is None:
        return ""

    text = str(value).replace("\xa0", " ")
    return " ".join(text.split())


def is_section_header(label: str) -> bool:
    return bool(SECTION_HEADER_RE.match(label))


def is_total_label(label: str) -> bool:
    return bool(TOTAL_LABEL_RE.search(label))


def convert_insurer_wide_sheet_to_csv_and_json(
    excel_path: Path | str,
    csv_path: Path | str,
    json_path: Path | str,
    expected_title: str,
    expected_data_column_count: int,
) -> dict:
    source_path = Path(excel_path)
    target_csv_path = Path(csv_path)
    target_json_path = Path(json_path)

    workbook = load_workbook(source_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    _validate_insurer_wide_layout(worksheet, expected_title)

    merged_header_values = build_merged_value_lookup(
        worksheet,
        row_start=4,
        row_end=8,
        column_start=1,
        column_end=worksheet.max_column,
    )

    data_column_indexes = find_metric_column_indexes(
        worksheet,
        header_row_start=4,
        header_row_end=8,
        metric_start_column=3,
        merged_values=merged_header_values,
    )
    if len(data_column_indexes) != expected_data_column_count:
        raise ValueError(
            f"Expected {expected_data_column_count} data columns, got {len(data_column_indexes)}."
        )

    column_labels = extract_metric_labels_for_columns(
        worksheet,
        column_indexes=data_column_indexes,
        header_row_start=4,
        header_row_end=8,
        merged_values=merged_header_values,
    )
    report_metadata = _extract_insurer_wide_report_metadata(worksheet)
    id_column_labels = [normalize_text(worksheet.cell(4, column_index).value) for column_index in (1, 2)]

    insurer_rows: list[dict] = []
    summary_row: dict | None = None
    footnote_rows: list[dict] = []

    for row_index in range(9, worksheet.max_row + 1):
        first_cell = normalize_text(worksheet.cell(row_index, 1).value)
        second_cell = normalize_text(worksheet.cell(row_index, 2).value)
        values = [worksheet.cell(row_index, column_index).value for column_index in data_column_indexes]
        has_data = any(value is not None for value in values)

        if not first_cell and not second_cell and not has_data:
            continue

        if first_cell.startswith("*"):
            footnote_rows.append(
                {
                    "excel_row": row_index,
                    "text": first_cell,
                }
            )
            continue

        if first_cell == "ИТОГО:":
            summary_row = {
                "excel_row": row_index,
                "label": first_cell,
                "values": {
                    column_label: value
                    for column_label, value in zip(column_labels, values, strict=True)
                },
            }
            continue

        if not first_cell or not second_cell:
            raise ValueError(f"Expected insurer identifiers in row {row_index}.")
        if not has_data:
            raise ValueError(f"Expected data values in row {row_index}.")

        insurer_rows.append(
            {
                "excel_row": row_index,
                id_column_labels[0]: first_cell,
                id_column_labels[1]: second_cell,
                "values": values,
            }
        )

    if summary_row is None:
        raise ValueError("Summary row was not found.")

    skipped_empty_columns = [
        {"excel_column": get_column_letter(column_index)}
        for column_index in range(data_column_indexes[-1] + 1, worksheet.max_column + 1)
    ]

    target_csv_path.parent.mkdir(parents=True, exist_ok=True)
    target_json_path.parent.mkdir(parents=True, exist_ok=True)

    with target_csv_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)
        writer.writerow([*id_column_labels, *column_labels])

        for row in insurer_rows:
            writer.writerow([row[id_column_labels[0]], row[id_column_labels[1]], *row["values"]])

    metadata = {
        "source_file": source_path.name,
        "sheet_name": worksheet.title,
        "report_date": report_metadata["report_date"],
        "report_date_iso": report_metadata["report_date_iso"],
        "report_period": report_metadata["report_period"],
        "hierarchy_separator": HIERARCHY_SEPARATOR,
        "id_columns": [
            {"excel_column": get_column_letter(column_index), "label": label}
            for column_index, label in zip((1, 2), id_column_labels, strict=True)
        ],
        "insurer_row_count": len(insurer_rows),
        "insurance_columns": [
            {
                "excel_column": get_column_letter(column_index),
                "path": column_label.split(HIERARCHY_SEPARATOR),
                "column_name": column_label,
            }
            for column_index, column_label in zip(data_column_indexes, column_labels, strict=True)
        ],
        "summary_row": summary_row,
        "footnote_rows": footnote_rows,
        "skipped_empty_columns": skipped_empty_columns,
    }

    with target_json_path.open("w", encoding="utf-8") as file:
        json.dump(metadata, file, ensure_ascii=False, indent=2)
        file.write("\n")

    return metadata


def _validate_insurer_wide_layout(worksheet, expected_title: str) -> None:
    if normalize_text(worksheet["A1"].value) != expected_title:
        raise ValueError("Unexpected label in A1.")

    first_id_label = normalize_text(worksheet["A4"].value)
    second_id_label = normalize_text(worksheet["B4"].value)

    if first_id_label != "Регистрационный номер записи страховщика в едином государственном реестре субъектов страхового дела":
        raise ValueError("Unexpected label in A4.")
    if second_id_label != "Полное наименование страховщика":
        raise ValueError("Unexpected label in B4.")


def _extract_insurer_wide_report_metadata(worksheet) -> dict[str, str]:
    report_date_cell = normalize_text(worksheet["A2"].value)
    report_period_cell = normalize_text(worksheet["A3"].value)

    if not report_date_cell.startswith("Дата составления отчета: "):
        raise ValueError("Unexpected report date cell format in A2.")
    if not report_period_cell.startswith("Отчетный период: "):
        raise ValueError("Unexpected report period cell format in A3.")

    report_date = report_date_cell.removeprefix("Дата составления отчета: ").strip()
    report_period = report_period_cell.removeprefix("Отчетный период: ").strip()
    report_date_iso = datetime.strptime(report_date, "%d.%m.%Y").date().isoformat()

    return {
        "report_date": report_date,
        "report_date_iso": report_date_iso,
        "report_period": report_period,
    }
