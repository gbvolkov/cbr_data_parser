from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from cbr_data_parser.utils.cbr_parser_utils import (
    HIERARCHY_SEPARATOR,
    build_merged_value_lookup,
    extract_metric_labels_for_columns,
    find_metric_column_indexes,
    normalize_text,
)


HEADER_ROW_START = 5
HEADER_ROW_END = 9
ROW_LABEL_COLUMN = 1
METRIC_START_COLUMN = 2
SUMMARY_ROW_INDEX = 10
ROW_LABEL_HEADER = "Наименования федеральных округов и субъектов Российской Федерации"


def convert_regional_wide_sheet_to_csv_and_json(
    excel_path: Path | str,
    csv_path: Path | str,
    json_path: Path | str,
    expected_title: str,
    expected_data_column_count: int,
) -> dict:
    source_path = Path(excel_path)
    target_csv_path = Path(csv_path)
    target_json_path = Path(json_path)

    workbook = load_workbook(source_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    _validate_layout(worksheet, expected_title)

    merged_header_values = build_merged_value_lookup(
        worksheet,
        row_start=HEADER_ROW_START,
        row_end=HEADER_ROW_END,
        column_start=METRIC_START_COLUMN,
        column_end=worksheet.max_column,
    )

    metric_column_indexes = find_metric_column_indexes(
        worksheet,
        header_row_start=HEADER_ROW_START,
        header_row_end=HEADER_ROW_END,
        metric_start_column=METRIC_START_COLUMN,
        merged_values=merged_header_values,
    )
    if len(metric_column_indexes) != expected_data_column_count:
        raise ValueError(
            f"Expected {expected_data_column_count} data columns, got {len(metric_column_indexes)}."
        )

    metric_labels = extract_metric_labels_for_columns(
        worksheet,
        column_indexes=metric_column_indexes,
        header_row_start=HEADER_ROW_START,
        header_row_end=HEADER_ROW_END,
        merged_values=merged_header_values,
    )
    duplicate_metric_labels = sorted(
        {
            metric_label
            for metric_label in metric_labels
            if metric_labels.count(metric_label) > 1
        }
    )
    if duplicate_metric_labels:
        raise ValueError(f"Duplicate metric labels found: {duplicate_metric_labels}")

    report_metadata = _extract_report_metadata(worksheet)
    row_label_header = normalize_text(worksheet.cell(HEADER_ROW_START, ROW_LABEL_COLUMN).value)
    if row_label_header != ROW_LABEL_HEADER:
        raise ValueError("Unexpected label in A5.")

    region_rows: list[dict] = []
    summary_row: dict | None = None
    footnote_rows: list[dict] = []

    for row_index in range(SUMMARY_ROW_INDEX, worksheet.max_row + 1):
        row_label = normalize_text(worksheet.cell(row_index, ROW_LABEL_COLUMN).value)
        values = [worksheet.cell(row_index, column_index).value for column_index in metric_column_indexes]
        has_data = any(value is not None for value in values)

        if not row_label and not has_data:
            continue

        if row_label.startswith("*"):
            footnote_rows.append(
                {
                    "excel_row": row_index,
                    "text": row_label,
                }
            )
            continue

        if row_label.upper().rstrip(":") == "ИТОГО":
            summary_row = {
                "excel_row": row_index,
                "label": row_label,
                "values": {
                    metric_label: value
                    for metric_label, value in zip(metric_labels, values, strict=True)
                },
            }
            continue

        if not row_label:
            raise ValueError(f"Expected region label in row {row_index}.")
        if not has_data:
            raise ValueError(f"Expected data values in row {row_index}.")

        region_rows.append(
            {
                "excel_row": row_index,
                "label": row_label,
                "values": values,
            }
        )

    if summary_row is None:
        raise ValueError("Summary row was not found.")

    skipped_empty_columns = [
        {"excel_column": get_column_letter(column_index)}
        for column_index in range(metric_column_indexes[-1] + 1, worksheet.max_column + 1)
    ]

    target_csv_path.parent.mkdir(parents=True, exist_ok=True)
    target_json_path.parent.mkdir(parents=True, exist_ok=True)

    with target_csv_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)
        writer.writerow([row_label_header, *metric_labels])

        for row in region_rows:
            writer.writerow([row["label"], *row["values"]])

    metadata = {
        "source_file": source_path.name,
        "sheet_name": worksheet.title,
        "sheet_description": normalize_text(worksheet["A1"].value),
        "report_date": report_metadata["report_date"],
        "report_date_iso": report_metadata["report_date_iso"],
        "report_period": report_metadata["report_period"],
        "row_label_column": {
            "excel_column": get_column_letter(ROW_LABEL_COLUMN),
            "label": row_label_header,
        },
        "hierarchy_separator": HIERARCHY_SEPARATOR,
        "region_row_count": len(region_rows),
        "metric_columns": [
            {
                "excel_column": get_column_letter(column_index),
                "path": metric_label.split(HIERARCHY_SEPARATOR),
                "column_name": metric_label,
            }
            for column_index, metric_label in zip(metric_column_indexes, metric_labels, strict=True)
        ],
        "summary_row": summary_row,
        "footnote_rows": footnote_rows,
        "skipped_empty_columns": skipped_empty_columns,
    }

    with target_json_path.open("w", encoding="utf-8") as file:
        json.dump(metadata, file, ensure_ascii=False, indent=2)
        file.write("\n")

    return metadata


def _validate_layout(worksheet, expected_title: str) -> None:
    if normalize_text(worksheet["A1"].value) != expected_title:
        raise ValueError("Unexpected label in A1.")


def _extract_report_metadata(worksheet) -> dict[str, str]:
    report_date_cell = normalize_text(worksheet["A3"].value)
    report_period_cell = normalize_text(worksheet["A4"].value)

    if not report_date_cell.startswith("Дата составления отчета: "):
        raise ValueError("Unexpected report date cell format in A3.")
    if not report_period_cell.startswith("Отчетный период: "):
        raise ValueError("Unexpected report period cell format in A4.")

    report_date = report_date_cell.removeprefix("Дата составления отчета: ").strip()
    report_period = report_period_cell.removeprefix("Отчетный период: ").strip()
    report_date_iso = datetime.strptime(report_date, "%d.%m.%Y").date().isoformat()

    return {
        "report_date": report_date,
        "report_date_iso": report_date_iso,
        "report_period": report_period,
    }
