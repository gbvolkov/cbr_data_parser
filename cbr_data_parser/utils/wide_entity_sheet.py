from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from cbr_data_parser.utils.cbr_parser_utils import (
    HIERARCHY_SEPARATOR,
    build_merged_value_lookup,
    extract_metric_labels_for_columns,
    find_metric_column_indexes,
    normalize_text,
)


def convert_wide_entity_sheet_to_csv_and_json(
    excel_path: Path | str,
    csv_path: Path | str,
    json_path: Path | str,
    *,
    expected_title: str,
    expected_id_label_a: str,
    expected_id_label_b: str,
    header_row_start: int,
    header_row_end: int,
    summary_row_index: int,
    summary_label: str,
    data_start_row: int,
    expected_metric_count: int,
    metric_start_column: int = 3,
) -> dict:
    source_path = Path(excel_path)
    target_csv_path = Path(csv_path)
    target_json_path = Path(json_path)

    workbook = load_workbook(source_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    _validate_layout(
        worksheet,
        expected_title=expected_title,
        expected_id_label_a=expected_id_label_a,
        expected_id_label_b=expected_id_label_b,
    )

    merged_header_values = build_merged_value_lookup(
        worksheet,
        row_start=header_row_start,
        row_end=header_row_end,
        column_start=1,
        column_end=worksheet.max_column,
    )

    metric_column_indexes = find_metric_column_indexes(
        worksheet,
        header_row_start=header_row_start,
        header_row_end=header_row_end,
        metric_start_column=metric_start_column,
        merged_values=merged_header_values,
    )
    if len(metric_column_indexes) != expected_metric_count:
        raise ValueError(
            f"Expected {expected_metric_count} metric columns, got {len(metric_column_indexes)}."
        )

    metric_labels = extract_metric_labels_for_columns(
        worksheet,
        column_indexes=metric_column_indexes,
        header_row_start=header_row_start,
        header_row_end=header_row_end,
        merged_values=merged_header_values,
    )
    report_metadata = _extract_report_metadata(worksheet)
    id_column_labels = [
        normalize_text(worksheet.cell(header_row_start, column_index).value)
        for column_index in (1, 2)
    ]

    entity_rows: list[dict] = []
    summary_row: dict | None = None
    footnote_rows: list[dict] = []

    for row_index in range(summary_row_index, worksheet.max_row + 1):
        first_cell = normalize_text(worksheet.cell(row_index, 1).value)
        second_cell = normalize_text(worksheet.cell(row_index, 2).value)
        values = [worksheet.cell(row_index, column_index).value for column_index in metric_column_indexes]
        has_data = any(value is not None for value in values)

        if row_index < data_start_row and row_index != summary_row_index:
            continue

        if not first_cell and not second_cell and not has_data:
            continue

        if first_cell.startswith("*"):
            footnote_rows.append(
                {
                    "excel_row": row_index,
                    "text": first_cell,
                }
            )
            continue

        if row_index == summary_row_index or first_cell == summary_label:
            summary_row = {
                "excel_row": row_index,
                "label": first_cell,
                "values": {
                    metric_label: value for metric_label, value in zip(metric_labels, values, strict=True)
                },
            }
            continue

        if not first_cell or not second_cell:
            raise ValueError(f"Expected entity identifiers in row {row_index}.")
        if not has_data:
            raise ValueError(f"Expected metric values in row {row_index}.")

        entity_rows.append(
            {
                "excel_row": row_index,
                id_column_labels[0]: first_cell,
                id_column_labels[1]: second_cell,
                "values": values,
            }
        )

    if summary_row is None:
        raise ValueError("Summary row was not found.")

    skipped_empty_columns = [
        {"excel_column": get_column_letter(column_index)}
        for column_index in range(metric_column_indexes[-1] + 1, worksheet.max_column + 1)
    ]

    target_csv_path.parent.mkdir(parents=True, exist_ok=True)
    target_json_path.parent.mkdir(parents=True, exist_ok=True)

    with target_csv_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)
        writer.writerow([*id_column_labels, *metric_labels])

        for row in entity_rows:
            writer.writerow([row[id_column_labels[0]], row[id_column_labels[1]], *row["values"]])

    metadata = {
        "source_file": source_path.name,
        "sheet_name": worksheet.title,
        "sheet_description": normalize_text(worksheet["A1"].value),
        "report_date": report_metadata["report_date"],
        "report_date_iso": report_metadata["report_date_iso"],
        "report_period": report_metadata["report_period"],
        "hierarchy_separator": HIERARCHY_SEPARATOR,
        "id_columns": [
            {"excel_column": get_column_letter(column_index), "label": label}
            for column_index, label in zip((1, 2), id_column_labels, strict=True)
        ],
        "record_row_count": len(entity_rows),
        "metric_columns": [
            {
                "excel_column": get_column_letter(column_index),
                "path": metric_label.split(HIERARCHY_SEPARATOR),
                "column_name": metric_label,
            }
            for column_index, metric_label in zip(metric_column_indexes, metric_labels, strict=True)
        ],
        "summary_row": summary_row,
        "footnote_rows": footnote_rows,
        "skipped_empty_columns": skipped_empty_columns,
    }

    with target_json_path.open("w", encoding="utf-8") as file:
        json.dump(metadata, file, ensure_ascii=False, indent=2)
        file.write("\n")

    return metadata


def _validate_layout(worksheet, expected_title: str, expected_id_label_a: str, expected_id_label_b: str) -> None:
    if normalize_text(worksheet["A1"].value) != expected_title:
        raise ValueError("Unexpected label in A1.")

    if normalize_text(worksheet["A4"].value) != expected_id_label_a:
        raise ValueError("Unexpected label in A4.")
    if normalize_text(worksheet["B4"].value) != expected_id_label_b:
        raise ValueError("Unexpected label in B4.")


def _extract_report_metadata(worksheet) -> dict[str, str]:
    report_date_cell = normalize_text(worksheet["A2"].value)
    report_period_cell = normalize_text(worksheet["A3"].value)

    if not report_date_cell.startswith("Дата составления отчета: "):
        raise ValueError("Unexpected report date cell format in A2.")
    if not report_period_cell.startswith("Отчетный период: "):
        raise ValueError("Unexpected report period cell format in A3.")

    report_date = report_date_cell.removeprefix("Дата составления отчета: ").strip()
    report_period = report_period_cell.removeprefix("Отчетный период: ").strip()
    report_date_iso = datetime.strptime(report_date, "%d.%m.%Y").date().isoformat()

    return {
        "report_date": report_date,
        "report_date_iso": report_date_iso,
        "report_period": report_period,
    }
