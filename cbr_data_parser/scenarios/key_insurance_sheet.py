from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from cbr_data_parser.utils import (
    HIERARCHY_SEPARATOR,
    METRIC_END_COLUMN,
    METRIC_START_COLUMN,
    build_merged_header_lookup,
    extract_insurance_columns,
    extract_metric_labels,
    extract_report_metadata,
    find_data_row_indexes,
    validate_layout,
)


def convert_key_insurance_sheet_to_csv_and_json(
    excel_path: Path | str,
    csv_path: Path | str,
    json_path: Path | str,
) -> dict:
    source_path = Path(excel_path)
    target_csv_path = Path(csv_path)
    target_json_path = Path(json_path)

    workbook = load_workbook(source_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    validate_layout(worksheet)

    merged_header_values = build_merged_header_lookup(worksheet)
    report_metadata = extract_report_metadata(worksheet)
    metric_labels = extract_metric_labels(worksheet, merged_header_values)
    data_row_indexes = find_data_row_indexes(worksheet)
    insurance_columns = extract_insurance_columns(worksheet, data_row_indexes)

    target_csv_path.parent.mkdir(parents=True, exist_ok=True)
    target_json_path.parent.mkdir(parents=True, exist_ok=True)

    with target_csv_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["Показатель", *[column["column_name"] for column in insurance_columns]])

        for metric_index, metric_label in enumerate(metric_labels):
            writer.writerow(
                [metric_label, *[column["values"][metric_index] for column in insurance_columns]]
            )

    metadata = {
        "source_file": source_path.name,
        "sheet_name": worksheet.title,
        "report_date": report_metadata["report_date"],
        "report_date_iso": report_metadata["report_date_iso"],
        "report_period": report_metadata["report_period"],
        "row_id_column": "Показатель",
        "hierarchy_separator": HIERARCHY_SEPARATOR,
        "metric_rows": metric_labels,
        "metric_sources": [
            {
                "excel_column": get_column_letter(column_index),
                "label": metric_label,
            }
            for column_index, metric_label in zip(
                range(METRIC_START_COLUMN, METRIC_END_COLUMN + 1), metric_labels, strict=True
            )
        ],
        "insurance_columns": [
            {
                "excel_row": column["excel_row"],
                "header_cell_column": get_column_letter(column["header_cell_column"]),
                "source_label": column["source_label"],
                "path": column["path"],
                "column_name": column["column_name"],
            }
            for column in insurance_columns
        ],
    }

    with target_json_path.open("w", encoding="utf-8") as file:
        json.dump(metadata, file, ensure_ascii=False, indent=2)
        file.write("\n")

    return metadata
