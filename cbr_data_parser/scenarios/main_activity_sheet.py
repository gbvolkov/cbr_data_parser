from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from cbr_data_parser.utils import (
    HIERARCHY_SEPARATOR,
    build_merged_value_lookup,
    extract_metric_labels_for_columns,
    extract_report_metadata,
    extract_row_label_for_columns,
    extract_row_path,
    find_metric_column_indexes,
    normalize_text,
    row_has_any_metric_value,
)


HEADER_ROW_START = 5
HEADER_ROW_END = 7
ROW_HEADER_START_COLUMN = 1
ROW_HEADER_END_COLUMN = 5
METRIC_START_COLUMN = 6
EXPECTED_METRIC_COUNT = 37
DATA_ROW_START = 8


def convert_main_activity_sheet_to_csv_and_json(
    excel_path: Path | str,
    csv_path: Path | str,
    json_path: Path | str,
) -> dict:
    source_path = Path(excel_path)
    target_csv_path = Path(csv_path)
    target_json_path = Path(json_path)

    workbook = load_workbook(source_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    _validate_layout(worksheet)

    merged_header_values = build_merged_value_lookup(
        worksheet,
        row_start=HEADER_ROW_START,
        row_end=HEADER_ROW_END,
        column_start=METRIC_START_COLUMN,
    )
    row_header_values = build_merged_value_lookup(
        worksheet,
        row_start=DATA_ROW_START,
        row_end=worksheet.max_row,
        column_start=ROW_HEADER_START_COLUMN,
        column_end=ROW_HEADER_END_COLUMN,
    )

    metric_column_indexes = find_metric_column_indexes(
        worksheet,
        header_row_start=HEADER_ROW_START,
        header_row_end=HEADER_ROW_END,
        metric_start_column=METRIC_START_COLUMN,
        merged_values=merged_header_values,
    )
    if len(metric_column_indexes) != EXPECTED_METRIC_COUNT:
        raise ValueError(
            f"Expected {EXPECTED_METRIC_COUNT} metric columns, got {len(metric_column_indexes)}."
        )

    metric_labels = extract_metric_labels_for_columns(
        worksheet,
        column_indexes=metric_column_indexes,
        header_row_start=HEADER_ROW_START,
        header_row_end=HEADER_ROW_END,
        merged_values=merged_header_values,
    )
    report_metadata = extract_report_metadata(worksheet)

    insurance_columns: list[dict] = []
    skipped_empty_rows: list[dict] = []
    footnote_rows: list[dict] = []

    for row_index in range(DATA_ROW_START, worksheet.max_row + 1):
        path = extract_row_path(
            worksheet,
            row_index=row_index,
            header_start_column=ROW_HEADER_START_COLUMN,
            header_end_column=ROW_HEADER_END_COLUMN,
            merged_values=row_header_values,
        )
        if not path:
            continue

        header_cell_column, source_label = extract_row_label_for_columns(
            worksheet,
            row_index=row_index,
            header_start_column=ROW_HEADER_START_COLUMN,
            header_end_column=ROW_HEADER_END_COLUMN,
        )

        row_info = {
            "excel_row": row_index,
            "header_cell_column": get_column_letter(header_cell_column),
            "source_label": normalize_text(source_label),
            "path": path,
            "column_name": HIERARCHY_SEPARATOR.join(path),
        }

        if source_label.startswith("*"):
            footnote_rows.append(row_info)
            continue

        if not row_has_any_metric_value(worksheet, row_index, metric_column_indexes):
            skipped_empty_rows.append(row_info)
            continue

        values = [worksheet.cell(row_index, column_index).value for column_index in metric_column_indexes]
        insurance_columns.append({**row_info, "values": values})

    target_csv_path.parent.mkdir(parents=True, exist_ok=True)
    target_json_path.parent.mkdir(parents=True, exist_ok=True)

    with target_csv_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["Показатель", *[column["column_name"] for column in insurance_columns]])

        for metric_index, metric_label in enumerate(metric_labels):
            writer.writerow(
                [metric_label, *[column["values"][metric_index] for column in insurance_columns]]
            )

    metadata = {
        "source_file": source_path.name,
        "sheet_name": worksheet.title,
        "sheet_description": normalize_text(worksheet["A1"].value),
        "report_date": report_metadata["report_date"],
        "report_date_iso": report_metadata["report_date_iso"],
        "report_period": report_metadata["report_period"],
        "row_id_column": "Показатель",
        "hierarchy_separator": HIERARCHY_SEPARATOR,
        "metric_rows": metric_labels,
        "metric_sources": [
            {
                "excel_column": get_column_letter(column_index),
                "label": metric_label,
            }
            for column_index, metric_label in zip(metric_column_indexes, metric_labels, strict=True)
        ],
        "insurance_columns": [
            {
                "excel_row": column["excel_row"],
                "header_cell_column": column["header_cell_column"],
                "source_label": column["source_label"],
                "path": column["path"],
                "column_name": column["column_name"],
            }
            for column in insurance_columns
        ],
        "skipped_empty_rows": skipped_empty_rows,
        "footnote_rows": footnote_rows,
    }

    with target_json_path.open("w", encoding="utf-8") as file:
        json.dump(metadata, file, ensure_ascii=False, indent=2)
        file.write("\n")

    return metadata


def _validate_layout(worksheet) -> None:
    if normalize_text(worksheet["A5"].value) != "Вид страхования":
        raise ValueError("Expected cell A5 to contain 'Вид страхования'.")
